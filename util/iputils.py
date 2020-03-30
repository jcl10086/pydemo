#!/usr/bin/env python 
# -*- coding:utf-8 -*-

import requests
from openpyxl import load_workbook

def transfer_ip(ip):
    url = "http://ip-api.com/json/" + ip
    json = requests.get(url).json()
    if 'lat' in json and 'lon' in json:
        print ip + "\t" + str(json['lat']) + "\t" + str(json['lon'])

if __name__ == '__main__':
    workbook = load_workbook('d:\\1.xlsx')
    booksheet = workbook.active

    # 获取sheet页的行数据
    rows = booksheet.rows
    # 获取sheet页的列数据
    columns = booksheet.columns

    i = 0
    for row in rows:
        i = i + 1
        x = booksheet.cell(row=i, column=1).value
        # y = booksheet.cell(row=i, column=2).value
        transfer_ip(x)

