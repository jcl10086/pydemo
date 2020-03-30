#!/usr/bin/env python 
# -*- coding:utf-8 -*-

import requests
from openpyxl import load_workbook


workbook = load_workbook('d:\\abc.xlsx')
booksheet = workbook.active

#获取sheet页的行数据
rows = booksheet.rows
#获取sheet页的列数据
columns = booksheet.columns

i = 0
for row in rows:
    i = i+1
    x = booksheet.cell(row=i,column=1).value
    y = booksheet.cell(row=i, column=2).value

    if( x == 'X'):
        continue
    # print str(x) + "==" + str(y)
    url = "http://api.map.baidu.com/geoconv/v1/?coords=" + str(x) + "," + str(y) +"&from=1&to=5&ak=Gym17f1IOgTmTTVGinR5QdEnPHtkDCC6"
    json = requests.get(url).json()
    print str(json['result'][0]['x']) + "\t" + str(json['result'][0]['y'])

